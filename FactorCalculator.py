import openpyxl

COVERAGE_FIRST_COLUMN = 3
COVERAGE_ROW = 0
FACTOR_FIRST_ROW = 3
TERM_COLUMN = 0
AGE_COLUMN = 1
HEALTH_CLASS = ["Preferred Plus", "Preferred", "Standard Plus", "Standard"]
NUMBER_OF_HEALTH_CLASS = len(HEALTH_CLASS)
FACTOR_FIRST_COLUMN = 3


def create_coverage_ranges(work_sheet):
    """
    :param work_sheet: The rates-table sheet file.
    :return: The cover ranges like - [("100", "249"),....]
    """
    coverage_ranges = []
    for col in work_sheet.iter_cols(COVERAGE_FIRST_COLUMN, work_sheet.max_column):
        coverage_range_val = col[COVERAGE_ROW].value
        if coverage_range_val is not None:
            coverage_ranges.append(create_range_from_excel_val(coverage_range_val))
    return coverage_ranges


def create_range_from_excel_val(coverage_range_val):
    """
    :param coverage_range_val: Range string val Like "$100k - $249k"
    :return: range tuple like - ("100", "249")
    """
    return tuple(coverage_range_val[1:-1].split("k - $"))


class FactorCalculator:
    ERROR_CODE = -1

    def __init__(self, rates_table_excel_path):
        workbook = openpyxl.load_workbook(rates_table_excel_path)
        work_sheet = workbook.active
        self.__coverage_ranges = create_coverage_ranges(work_sheet)
        self.__factor_rates_dict = self.__create_factor_rates_dict(work_sheet)

    def __create_factor_rates_dict(self, work_sheet):
        """
        :param work_sheet: The rates-table sheet file.
        :return: Factor rates dict like - {term: {age: {coverage rang: {health-class: factor, ...},...}, ...}, ...}
        """
        factor_rates_dict = {}
        for row_index in range(FACTOR_FIRST_ROW, work_sheet.max_row):
            if work_sheet[row_index][0].value is None:
                break
            term = work_sheet[row_index][TERM_COLUMN].value
            age = work_sheet[row_index][AGE_COLUMN].value
            factors = {}
            # Crate {coverage range: {health-class: factor}} dict
            for i, coverage_range in enumerate(self.__coverage_ranges):
                factors_for_health_class_dict = {}
                first_col_index = FACTOR_FIRST_COLUMN + i * NUMBER_OF_HEALTH_CLASS
                # Crate {health-class: factor} dict
                for j, col in enumerate(work_sheet.iter_cols(first_col_index,
                                                             first_col_index + NUMBER_OF_HEALTH_CLASS - 1)):
                    factors_for_health_class_dict[HEALTH_CLASS[j]] = col[row_index - 1].value

                factors[coverage_range] = factors_for_health_class_dict
            if term not in factor_rates_dict:
                factor_rates_dict[term] = {age: factors}
            else:
                factor_rates_dict[term][age] = factors
        return factor_rates_dict

    def calculate_factor(self, health_class, term, age, coverage):
        if health_class not in HEALTH_CLASS:
            return FactorCalculator.ERROR_CODE
        coverage_ranges_dict = self.__factor_rates_dict[term][age]
        for coverage_range in coverage_ranges_dict.keys():
            if int(coverage_range[0]) <= coverage/1000 <= int(coverage_range[1]):
                return coverage_ranges_dict[coverage_range][health_class]
        return FactorCalculator.ERROR_CODE
