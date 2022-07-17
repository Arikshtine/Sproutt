import openpyxl

HEALTH_CLASS_ROW = 3
HEALTH_CLASS_FIRST_COLUMN = 3
HEIGHT_FIRST_ROW = 6
FEET_COLUMN = 0
INCH_COLUMN = 1
WEIGHT_FIRST_COLUMN = 3
DECLINED_CODE_NAME = "Declined"


def create_health_class_list(work_sheet):
    """
    :param work_sheet: The health class sheet file.
    :return: List of all health class.
    """
    return [col[HEALTH_CLASS_ROW].value for col in work_sheet.iter_cols(HEALTH_CLASS_FIRST_COLUMN,
                                                                        work_sheet.max_column)]


def create_health_class_dict(work_sheet):
    """
    :param work_sheet: The health class sheet file.
    :return: Dict of {feet: {inch: [weight],...},...}
    """
    health_class_dict = {}
    for row_index in range(HEIGHT_FIRST_ROW, work_sheet.max_row):
        if work_sheet[row_index][0].value is None:
            break

        feet = ''.join(filter(str.isdigit, work_sheet[row_index][FEET_COLUMN].value))
        inch = ''.join(filter(str.isdigit, work_sheet[row_index][INCH_COLUMN].value))
        weights = [col[row_index - 1].value for col in work_sheet.iter_cols(WEIGHT_FIRST_COLUMN, work_sheet.max_column)]
        if feet not in health_class_dict:
            health_class_dict[feet] = {inch: weights}
        else:
            health_class_dict[feet][inch] = weights
    return health_class_dict


class HealthClassDecider:
    """
    An object that decides which health class you belong to.
    """

    def __init__(self, health_class_excel_path):
        workbook = openpyxl.load_workbook(health_class_excel_path)
        work_sheet = workbook.active
        self.__health_class_list = create_health_class_list(work_sheet)
        self.__health_class_dict = create_health_class_dict(work_sheet)

    def calculate_health_class(self, height, weight):
        feet, inch = height.split(" ft ")
        weights = self.__health_class_dict[feet][inch]
        if weight < weights[0]:
            return DECLINED_CODE_NAME
        for i in range(0, len(weights)-1):
            if weights[i] <= weight < weights[i+1]:
                return self.__health_class_list[i]
        return DECLINED_CODE_NAME
